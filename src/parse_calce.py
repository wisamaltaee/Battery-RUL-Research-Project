"""
parse_calce.py — Parse raw CALCE Arbin .xlsx logs (one battery = many dated
files, each containing a Channel_x-xxx sheet of row-level time series data)
into one unified, time-sorted DataFrame per cell.
"""

import os
import re
import pandas as pd
from openpyxl import load_workbook

COLS = [
    "Data_Point", "Test_Time(s)", "Date_Time", "Step_Time(s)", "Step_Index",
    "Cycle_Index", "Current(A)", "Voltage(V)", "Charge_Capacity(Ah)",
    "Discharge_Capacity(Ah)", "Charge_Energy(Wh)", "Discharge_Energy(Wh)",
    "dV/dt(V/s)", "Internal_Resistance(Ohm)",
]


def _read_one_file(path):
    wb = load_workbook(path, read_only=True)
    data_sheet = [s for s in wb.sheetnames if re.match(r"channel_\d", s, re.I)]
    if not data_sheet:
        return None
    ws = wb[data_sheet[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = list(header)
    col_idx = {c: header.index(c) for c in COLS if c in header}
    records = []
    for r in rows:
        records.append({c: r[i] for c, i in col_idx.items()})
    df = pd.DataFrame.from_records(records)
    return df


def parse_battery(folder: str) -> pd.DataFrame:
    """Read every .xlsx in `folder`, concatenate, sort by real elapsed time."""
    files = sorted(f for f in os.listdir(folder) if f.endswith(".xlsx"))
    frames = []
    for f in files:
        df = _read_one_file(os.path.join(folder, f))
        if df is not None and len(df):
            df["__source_file"] = f
            frames.append(df)
    full = pd.concat(frames, ignore_index=True)
    full = full.sort_values("Date_Time").reset_index(drop=True)

    # Cycle_Index resets to 1 in EVERY file (each file = new Arbin schedule
    # run), so raw Cycle_Index is not globally unique across the battery's
    # life. Build a global cycle counter from (file order, local cycle).
    file_order = {f: i for i, f in enumerate(files)}
    full["__file_order"] = full["__source_file"].map(file_order)
    full = full.sort_values(["__file_order", "Cycle_Index", "Data_Point"]).reset_index(drop=True)

    # global_cycle increments every time (file, local cycle) changes
    key = list(zip(full["__file_order"], full["Cycle_Index"]))
    global_cycle = [1]
    for i in range(1, len(key)):
        global_cycle.append(global_cycle[-1] + (1 if key[i] != key[i - 1] else 0))
    full["global_cycle"] = global_cycle
    return full


if __name__ == "__main__":
    import sys
    df = parse_battery(sys.argv[1] if len(sys.argv) > 1 else "data/raw/calce/CS2_35")
    print(df.shape)
    print(df[["Date_Time", "Cycle_Index", "global_cycle", "Current(A)", "Voltage(V)", "Discharge_Capacity(Ah)"]].head())
    print("Global cycles found:", df["global_cycle"].nunique())
